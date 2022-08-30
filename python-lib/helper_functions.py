# -*- coding: utf-8 -*-
import dataiku
import pandas as pd, numpy as np
from dataiku import pandasutils as pdu

from dataiku import SQLExecutor2

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import io 
import re

from tempfile import NamedTemporaryFile


# Write Excel to managed folder
def write_wb_to_managed_folder(wb,output_folder, file_name):
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = tmp.read()
        with output_folder.get_writer(file_name) as w:
            w.write(output)

# Read Input Template as Workbook
def read_wb_from_managed_folder(input_folder):
    template_file_name = input_folder.list_paths_in_partition()[0]
    with input_folder.get_download_stream(template_file_name) as f:
        bytes_in = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(bytes_in)
    return wb

# Write dataset to sheet at an Row/ Column position
def populate_table_in_ws(df, ws, start_row, start_col):
    df_np = df.values
    for row_num in range(df.shape[0]):
        for col_num in range (df.shape[1]):
            ws.cell(row = (row_num + start_row), column = (col_num +start_col)).value =  df_np[row_num][col_num]
    return ws


# Find a SQL tag in worksheets and return coordinates
def find_tags_in_ws(ws,query_tag,row_max, col_max):
    tags = []
    for row_idx in range(1,row_max):
        for col_idx in range (1,col_max):
            value = ws.cell(row = row_idx,column =col_idx).value
            if value != None:
                if (str(value).startswith(query_tag)):
                    tags.append([value,row_idx, col_idx])
    return tags


def get_df_from_query(cnx_name,query):
    executor = SQLExecutor2(connection=cnx_name)
    df = executor.query_to_df(query)
    return df

def get_project_var(project_var):
    client = dataiku.api_client()
    project_api = client.get_default_project()
    v = project_api.get_variables()
    var_value = v["standard"].get(project_var,None)
    return var_value

def insert_project_variables_in_query(query):
    #pattern = re.compile("@([a-zA-Z0-9_]*)[; ]?")
    pattern = re.compile("@([a-zA-Z0-9_]*)")
    variables = pattern.findall(query)
    query = query.replace("@","")
    for var in variables:
        var_value = get_project_var(var)
        if var_value !=None:
            query = query.replace(var, get_project_var(var))
        else:
            raise Exception("Project variable {} does not exist".format(var))
    return query

def poluate_wb_from_sql(wb,cnx_name,sql_tag,row_max=50, col_max=50):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tags = find_tags_in_ws(ws,sql_tag, row_max, col_max)
        for tag in tags:
            raw_query = tag[0]
            query = raw_query.replace(sql_tag,"")
            query = insert_project_variables_in_query(query)
            df = get_df_from_query(cnx_name,query)
            ws = populate_table_in_ws(df, ws,tag[1], tag[2])
            
def popluate_wb_from_dataset(wb, insert_tag, row_max=50, col_max=50):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tags = find_tags_in_ws(ws, insert_tag, row_max, col_max)
        for tag in tags:
            if len(tag)!=3:
                raise Exception("Tag cannot be properly extracted from sheet {}".format(sheet_name))
            if "." not in tag[0]:
                raise Exception("Tag {} is not well formatted".format(tag[0]))
            dataset_name = tag[0].split(".")[1]
            dataset = dataiku.Dataset(dataset_name)
            df = dataset.get_dataframe()
            populate_table_in_ws(df, ws, tag[1], tag[2])